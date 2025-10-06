
import logging
import json
from math import ceil
from datetime import datetime
from datetime import timedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection, transaction
from django.http import (
    JsonResponse,
    HttpResponseBadRequest,
    HttpResponseNotAllowed,
    HttpResponseForbidden,
)
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.http import urlencode
from django.views.decorators.http import require_GET, require_POST, require_http_methods

import mysql.connector
from mysql.connector import Error, IntegrityError
from django.http import JsonResponse, HttpResponseBadRequest
from django.db import connection
from datetime import datetime
from django.views.decorators.http import require_GET, require_POST
from django.http import JsonResponse
from django.db import connection, transaction
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from decimal import Decimal, ROUND_HALF_UP
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest
from django.db import connection, transaction
from django.template.loader import render_to_string
import json
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date, timedelta
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest
from django.db import connection, transaction

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from datetime import date  # Add this if missing
import io  # Add this if missing
import openpyxl
from xhtml2pdf import pisa
from openpyxl.utils import get_column_letter

PAGE_SIZE = 10
# -------------------------
# LDAP helpers (use your ldap_utils)
# -------------------------
# We expect these functions to be provided in accounts.ldap_utils and accept optional
# username_password_for_conn param so they can use session credentials.
try:
    from accounts.ldap_utils import get_user_entry_by_username, get_reportees_for_user_dn
except Exception:
    def get_user_entry_by_username(username, username_password_for_conn=None):
        logger.warning("ldap_utils.get_user_entry_by_username not available")
        return None

    def get_reportees_for_user_dn(user_dn, username_password_for_conn=None):
        logger.warning("ldap_utils.get_reportees_for_user_dn not available")
        return []

logger = logging.getLogger(__name__)

# Default hours available per employee per month (can be overridden in settings)
HOURS_AVAILABLE_PER_MONTH = float(getattr(settings, "HOURS_AVAILABLE_PER_MONTH", 183.75))

# -------------------------
# DB helpers
# -------------------------

def get_wbs_options_for_iom(iom_id):
    with connection.cursor() as cur:
        cur.execute("SELECT seller_wbs_cc, buyer_wbs_cc FROM prism_wbs WHERE iom_id=%s LIMIT 1", [iom_id])
        row = cur.fetchone()
    if not row:
        return []
    seller, buyer = row
    opts = []
    if seller: opts.append({"code": seller, "label": f"Seller WBS: {seller}"})
    if buyer: opts.append({"code": buyer, "label": f"Buyer WBS: {buyer}"})
    return opts

def dictfetchall(cursor):
    """Return all rows from a cursor as a list of dicts."""
    cols = [c[0] for c in cursor.description] if cursor.description else []
    return [dict(zip(cols, row)) for row in cursor.fetchall()]

def get_connection():
    dbs = settings.DATABASES.get("default", {})
    return mysql.connector.connect(
        host=dbs.get("HOST", "127.0.0.1") or "127.0.0.1",
        port=int(dbs.get("PORT", 3306) or 3306),
        user=dbs.get("USER", "root") or "",
        password=dbs.get("PASSWORD", "root") or "",
        database=dbs.get("NAME", "feasdb") or "",
        charset="utf8mb4",
        use_unicode=True,
    )
def get_month_start_and_end(year_month):
    # year_month is "YYYY-MM" or a date; returns (date_start, date_end)
    if isinstance(year_month, str) and "-" in year_month:
        dt = datetime.strptime(year_month + "-01", "%Y-%m-%d").date()
    elif isinstance(year_month, date):
        dt = year_month.replace(day=1)
    else:
        dt = date.today().replace(day=1)
    # compute end of month
    next_month = (dt.replace(day=28) + timedelta(days=4)).replace(day=1)
    last_day = next_month - timedelta(days=1)
    return (dt, last_day)

# -------------------------------------------------------------------
# 1. CENTRALIZED BILLING PERIOD SOURCE OF TRUTH
# -------------------------------------------------------------------
def get_billing_period(year: int, month: int):
    """Fetch billing cycle start_date and end_date from monthly_hours_limit.
       Fallback to calendar month if not defined."""
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE year = %s AND month = %s
                LIMIT 1
            """, [year, month])
            row = cur.fetchone()
            if row and row[0] and row[1]:
                return row[0], row[1]
    except Exception as e:
        logger.exception("Error reading billing period: %s", e)

    # fallback to calendar month
    start = date(year, month, 1)
    next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    end = next_month - timedelta(days=1)
    return start, end

def _get_billing_period_for_year_month(year: int, month: int):
    """
    Query monthly_hours_limit for the given year & month.
    If start_date and end_date exist (non-null), return (start_date, end_date) as date objects.
    Otherwise return the calendar month first..last day tuple.

    This function avoids calling get_month_start_and_end directly and ensures the
    canonical billing period (if present) is used.
    """
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE year=%s AND month=%s
                LIMIT 1
            """, [int(year), int(month)])
            row = cur.fetchone()
            if row:
                sd_raw, ed_raw = row[0], row[1]
                sd = None
                ed = None
                # Accept DB date objects or strings
                if sd_raw:
                    if isinstance(sd_raw, date):
                        sd = sd_raw
                    else:
                        try:
                            sd = datetime.strptime(str(sd_raw).split(" ")[0], "%Y-%m-%d").date()
                        except Exception:
                            sd = None
                if ed_raw:
                    if isinstance(ed_raw, date):
                        ed = ed_raw
                    else:
                        try:
                            ed = datetime.strptime(str(ed_raw).split(" ")[0], "%Y-%m-%d").date()
                        except Exception:
                            ed = None
                if sd and ed:
                    return sd, ed
    except Exception:
        logger.exception("_get_billing_period_for_year_month db error")
    # fallback to calendar month
    try:
        # reuse the simple calendar month computation already present in get_month_start_and_end
        if isinstance(year, int) and isinstance(month, int):
            start = date(year, month, 1)
            next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
            end = next_month - timedelta(days=1)
            return start, end
    except Exception:
        pass
    # as a final fallback, return today's month
    today = date.today()
    s = today.replace(day=1)
    nm = (s.replace(day=28) + timedelta(days=4)).replace(day=1)
    return s, (nm - timedelta(days=1))


def get_billing_period_for_date(punch_date: date):
    """Find which billing cycle a given date falls into."""
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE %s BETWEEN start_date AND end_date
                LIMIT 1
            """, [punch_date])
            row = cur.fetchone()
            if row and row[0] and row[1]:
                return row[0], row[1]
    except Exception:
        logger.warning("Date %s not found in billing cycle", punch_date)
    # fallback to that date's calendar month
    return get_billing_period(punch_date.year, punch_date.month)

def _find_billing_period_for_date(d: date):
    """
    Find a billing period (start_date, end_date) that contains the given date d by scanning
    monthly_hours_limit rows where start_date and end_date are not null. If found return that period.
    Otherwise fallback to the calendar month containing d.
    """
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date, year, month
                FROM monthly_hours_limit
                WHERE start_date IS NOT NULL AND end_date IS NOT NULL
                  AND %s BETWEEN start_date AND end_date
                LIMIT 1
            """, [d])
            row = cur.fetchone()
            if row:
                sd_raw, ed_raw = row[0], row[1]
                sd = sd_raw if isinstance(sd_raw, date) else datetime.strptime(str(sd_raw).split(" ")[0], "%Y-%m-%d").date()
                ed = ed_raw if isinstance(ed_raw, date) else datetime.strptime(str(ed_raw).split(" ")[0], "%Y-%m-%d").date()
                return sd, ed
    except Exception:
        logger.exception("_find_billing_period_for_date DB error")
    # fallback: return calendar month for the date d
    try:
        year_month = f"{d.year}-{str(d.month).zfill(2)}"
        return get_billing_period(int(d.year), int(d.month))
    except Exception:
        # safe final fallback: today calendar month
        s = d.replace(day=1)
        nm = (s.replace(day=28) + timedelta(days=4)).replace(day=1)
        return s, (nm - timedelta(days=1))



def month_day_to_week_number_for_period(d: date, period_start: date, period_end: date = None):
    """
    Map a date 'd' to a 1-based week number relative to a billing period that begins at 'period_start'.
    Weeks are contiguous 7-day buckets starting at period_start:
      days 0-6 -> week 1, 7-13 -> week 2, ...
    This function dynamically supports more than 4 weeks if the billing period spans >28 days.
    Returns an integer >=1.

    If period_end provided, we compute the total weeks for the billing period; caller can use that
    to display week blocks dynamically.
    """
    try:
        if not period_start:
            return 1
        delta_days = (d - period_start).days
        week = (delta_days // 7) + 1
        if week < 1:
            week = 1
        # if period_end provided, cap to total weeks in period
        if period_end:
            total_days = (period_end - period_start).days + 1
            total_weeks = int(ceil(total_days / 7.0))
            if week > total_weeks:
                week = total_weeks
        return week
    except Exception:
        # conservative fallback based on calendar day-of-month
        try:
            return min(((d.day - 1) // 7) + 1, 5)
        except Exception:
            return 1


def month_day_to_week_number(d):
    """
    Convert a date d (a datetime.date) to a month-week bucket 1..4.
    Uses the same logic: days 1-7 -> week 1, 8-14 -> week 2, 15-21 -> week 3, >=22 -> week 4.
    """
    return min(((d.day - 1) // 7) + 1, 4)

def _ensure_user_from_ldap(request, samaccountname):
    """
    Ensure a 'users' row exists for the given LDAP identifier (username or email).
    Returns the users.id (int) for the row (create if missing).

    Behavior:
      - If a users row already has ldap_id == samaccountname or username == samaccountname or email == samaccountname -> return it
      - Else insert a new users row:
          username = part before '@' if samaccountname looks like an email, else samaccountname
          ldap_id = samaccountname (store canonical identifier)
          email = samaccountname if it looks like an email, else NULL
    """
    if not samaccountname:
        return None

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # try to find existing by ldap_id, username or email
        cur.execute(
            "SELECT id, ldap_id, username, email FROM users WHERE ldap_id = %s OR username = %s OR email = %s LIMIT 1",
            (samaccountname, samaccountname, samaccountname)
        )
        row = cur.fetchone()
        if row:
            return row["id"]

        # Prepare insert values
        username_val = samaccountname
        email_val = None
        if "@" in samaccountname:
            # username part before @
            username_val = samaccountname.split("@", 1)[0]
            email_val = samaccountname

        ins = conn.cursor()
        try:
            ins.execute(
                "INSERT INTO users (username, ldap_id, email, created_at) VALUES (%s, %s, %s, CURRENT_TIMESTAMP)",
                (username_val, samaccountname, email_val)
            )
            conn.commit()
            new_id = ins.lastrowid
        finally:
            try:
                ins.close()
            except Exception:
                pass
        return new_id
    except Exception:
        logger.exception("Error in _ensure_user_from_ldap for identifier: %s", samaccountname)
        return None
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def _get_local_ldap_entry(identifier):
    """
    Look up the local ldap_directory table using email, username or cn.
    Returns a dict with keys (username, email, cn, title) or None.
    """
    if not identifier:
        return None
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT username, email, cn, title
            FROM ldap_directory
            WHERE email = %s OR username = %s OR cn = %s
            LIMIT 1
        """, (identifier, identifier, identifier))
        return cur.fetchone()
    except Exception:
        logger.exception("Error reading ldap_directory for %s", identifier)
        return None
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

def _fetch_users():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, username, email FROM users ORDER BY username LIMIT 500")
        return cur.fetchall()
    finally:
        cur.close(); conn.close()


def _fetch_project(project_id):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM projects WHERE id=%s LIMIT 1", (project_id,))
        return cur.fetchone()
    finally:
        cur.close(); conn.close()

# projects/views.py
from django.shortcuts import render

def project_list(request):
    """
    Return projects visible to the logged-in user:
      - projects where p.pdl_user_id == ldap_username (email)
      - OR projects linked (prism_wbs.project_id) where prism_wbs.creator matches converted CN

    The view returns projects (as before) for client-side pagination.
    """
    # Get session values
    ldap_username = request.session.get("ldap_username")  # expected to be email or identifier
    cn = request.session.get("cn")  # stored as "LASTNAME FirstName ..." (e.g. "DEO Sant Anurag")

    # convert cn (LastName + FirstName...) to creator format (FirstName ... LastName)
    creator_name = None
    try:
        if cn:
            parts = str(cn).strip().split()
            if len(parts) >= 2:
                # move first token (last name) to the end
                creator_name = " ".join(parts[1:] + parts[:1])
            else:
                creator_name = cn.strip()
    except Exception:
        creator_name = None

    # If neither ldap_username nor creator_name present, return empty list (no projects)
    if not ldap_username and not creator_name:
        return render(request, "projects/project_list.html", {"projects": []})

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    projects = []
    try:
        # Build a safe SQL that selects projects satisfying either condition.
        # Use parameter placeholders for both ldap_username and creator_name.
        # We use LEFT JOIN with prism_wbs and GROUP BY project to avoid duplicates.
        sql = """
            SELECT DISTINCT p.id, p.name, p.oem_name, p.description,
                   p.start_date, p.end_date, p.pdl_user_id, p.pdl_name,
                   p.pm_user_id, p.pm_name, p.created_at
            FROM projects p
            LEFT JOIN prism_wbs w ON w.project_id = p.id
            WHERE 1=0
        """
        params = []

        if ldap_username:
            sql += " OR (p.pdl_user_id = %s)"
            params.append(ldap_username)

        if creator_name:
            # match prism_wbs.creator exactly to converted creator name
            sql += " OR (w.creator = %s)"
            params.append(creator_name)

        sql += " ORDER BY p.created_at DESC"

        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []

        # normalize rows for JSON consumption (dates -> ISO)
        for r in rows:
            projects.append({
                "id": r.get("id"),
                "name": r.get("name") or "",
                "oem_name": r.get("oem_name") or "",
                "description": r.get("description") or "",
                "start_date": (r.get("start_date").isoformat() if r.get("start_date") else None),
                "end_date": (r.get("end_date").isoformat() if r.get("end_date") else None),
                "pdl_user_id": r.get("pdl_user_id") or "",
                "pdl_name": r.get("pdl_name") or "",
                "pm_user_id": r.get("pm_user_id") or "",
                "pm_name": r.get("pm_name") or "",
                "created_at": (r.get("created_at").isoformat() if r.get("created_at") else None),
            })
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    return render(request, "projects/project_list.html", {"projects": projects})

def _get_all_coes():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name FROM coes ORDER BY name")
        return cur.fetchall()
    finally:
        cur.close(); conn.close()

def _assign_coes_to_project(project_id, coe_ids):
    """
    Given project_id and iterable of coe_ids, insert into project_coes table.
    This function is idempotent: it skips existing mappings and inserts new ones.
    """
    if not coe_ids:
        return
    conn = get_connection()
    cur = conn.cursor()
    try:
        for cid in coe_ids:
            try:
                cur.execute("INSERT INTO project_coes (project_id, coe_id) VALUES (%s, %s)", (project_id, cid))
                # commit per batch later
            except IntegrityError:
                # mapping exists — ignore
                continue
        conn.commit()
    finally:
        cur.close(); conn.close()

def _replace_project_coes(project_id, coe_ids):
    """
    Replace mappings for project: delete all existing and insert provided list (idempotent).
    """
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM project_coes WHERE project_id=%s", (project_id,))
        if coe_ids:
            for cid in coe_ids:
                try:
                    cur.execute("INSERT INTO project_coes (project_id, coe_id) VALUES (%s, %s)", (project_id, cid))
                except IntegrityError:
                    continue
        conn.commit()
    finally:
        cur.close(); conn.close()

@require_POST
def delete_project(request, project_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM projects WHERE id=%s", (project_id,))
        conn.commit()
    finally:
        cur.close(); conn.close()
    return redirect(reverse("projects:list"))

@require_POST
def create_coe(request):
    name = (request.POST.get("name") or "").strip()
    leader_username = request.POST.get("leader_username") or None
    description = request.POST.get("description") or None

    if not name:
        return HttpResponseBadRequest("COE name required")

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM coes WHERE name = %s LIMIT 1", (name,))
        if cur.fetchone():
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "COE with this name already exists."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    leader_user_id = None
    if leader_username:
        leader_user_id = _ensure_user_from_ldap(request,leader_username)

    conn2 = get_connection()
    cur2 = conn2.cursor()
    try:
        try:
            cur2.execute("INSERT INTO coes (name, leader_user_id, description) VALUES (%s, %s, %s)",
                         (name, leader_user_id, description))
            conn2.commit()
        except IntegrityError as e:
            logger.warning("create_coe IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "COE insert failed (duplicate)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur2.close(); conn2.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_POST
def edit_coe(request, coe_id):
    name = (request.POST.get("name") or "").strip()
    leader_username = request.POST.get("leader_username") or None
    description = request.POST.get("description") or None

    if not name:
        return HttpResponseBadRequest("COE name required")

    leader_user_id = None
    if leader_username:
        leader_user_id = _ensure_user_from_ldap(request,leader_username)

    conn = get_connection()
    cur = conn.cursor()
    try:
        try:
            cur.execute("UPDATE coes SET name=%s, leader_user_id=%s, description=%s WHERE id=%s",
                        (name, leader_user_id, description, coe_id))
            conn.commit()
        except IntegrityError as e:
            logger.warning("edit_coe IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "COE update failed (duplicate or constraint)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_POST
def create_domain(request):
    name = (request.POST.get("name") or "").strip()
    coe_id = request.POST.get("coe_id") or None
    lead_username = request.POST.get("lead_username") or None

    if not name:
        return HttpResponseBadRequest("Domain name required")

    try:
        coe_id_int = int(coe_id) if coe_id else None
    except Exception:
        coe_id_int = None

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM domains WHERE coe_id = %s AND name = %s LIMIT 1", (coe_id_int, name))
        if cur.fetchone():
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Domain with this name already exists for the selected COE."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    lead_user_id = None
    if lead_username:
        lead_user_id = _ensure_user_from_ldap(request,lead_username)

    conn2 = get_connection()
    cur2 = conn2.cursor()
    try:
        try:
            cur2.execute("INSERT INTO domains (coe_id, name, lead_user_id) VALUES (%s, %s, %s)",
                         (coe_id_int if coe_id_int else None, name, lead_user_id))
            conn2.commit()
        except IntegrityError as e:
            logger.warning("create_domain IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Domain insert failed (duplicate)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur2.close(); conn2.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_POST
def edit_domain(request, domain_id):
    name = (request.POST.get("name") or "").strip()
    coe_id = request.POST.get("coe_id") or None
    lead_username = request.POST.get("lead_username") or None

    if not name:
        return HttpResponseBadRequest("Domain name required")

    try:
        coe_id_int = int(coe_id) if coe_id else None
    except Exception:
        coe_id_int = None

    lead_user_id = None
    if lead_username:
        lead_user_id = _ensure_user_from_ldap(request,lead_username)

    conn = get_connection()
    cur = conn.cursor()
    try:
        try:
            cur.execute("UPDATE domains SET coe_id=%s, name=%s, lead_user_id=%s WHERE id=%s",
                        (coe_id_int if coe_id_int else None, name, lead_user_id, domain_id))
            conn.commit()
        except IntegrityError as e:
            logger.warning("edit_domain IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Domain update failed (duplicate or constraint)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_GET
def ldap_search(request):
    """
    AJAX endpoint used by projects/actions.js autocomplete.

    - Expects query param 'q'
    - Requires minimum 3 characters to search (client enforces this too)
    - First looks up the local `ldap_directory` table (username, email, cn, title)
    - Returns JSON: {"results": [ {sAMAccountName, mail, cn, title}, ... ] }
    - If local table returns no rows, falls back to live LDAP via accounts.ldap_utils (if available)
    """
    q = (request.GET.get("q") or "").strip()
    if len(q) < 3:
        # Return empty results for short queries (client requires min 3 chars)
        return JsonResponse({"results": []})

    results = []
    try:
        # 1) Query local ldap_directory table (preferred)
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            like = f"%{q}%"
            cur.execute("""
                SELECT username AS sAMAccountName,
                       COALESCE(email, '') AS mail,
                       COALESCE(cn, username) AS cn,
                       COALESCE(title, '') AS title
                FROM ldap_directory
                WHERE username LIKE %s OR email LIKE %s OR cn LIKE %s
                ORDER BY username LIMIT 40
            """, (like, like, like))
            rows = cur.fetchall() or []
            for r in rows:
                results.append({
                    "sAMAccountName": r.get("sAMAccountName") or "",
                    "mail": r.get("mail") or "",
                    "cn": r.get("cn") or "",
                    "title": r.get("title") or "",
                })
            print("Results from local ldap_directory:", results)
        finally:
            try:
                cur.close()
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass

        # 2) If no local results, optionally fall back to live LDAP (keeps previous behavior)
        if not results:
            try:
                from accounts import ldap_utils
                username = request.session.get("ldap_username")
                password = request.session.get("ldap_password")
                # if no session creds, skip live LDAP
                if username and password:
                    conn_ldap = ldap_utils._get_ldap_connection(username, password)
                    base_dn = getattr(settings, "LDAP_BASE_DN", "")
                    conn_ldap.search(
                        search_base=base_dn,
                        search_filter=f"(|(sAMAccountName=*{q}*)(cn=*{q}*)(mail=*{q}*))",
                        search_scope='SUBTREE',
                        attributes=['sAMAccountName', 'mail', 'cn', 'title']
                    )
                    for e in conn_ldap.entries:
                        results.append({
                            "sAMAccountName": str(getattr(e, 'sAMAccountName', '')) or "",
                            "mail": str(getattr(e, 'mail', '')) or "",
                            "cn": str(getattr(e, 'cn', '')) or "",
                            "title": str(getattr(e, 'title', '')) or "",
                        })
                    print("Results from live LDAP:", results)
                    try:
                        conn_ldap.unbind()
                    except Exception:
                        pass
            except Exception as ex:
                logger.warning("Live LDAP fallback failed or not available: %s", ex)

    except Exception as ex:
        # In case of unexpected DB failure, log and return empty list (avoid breaking UI)
        logger.exception("ldap_search: unexpected error: %s", ex)
        return JsonResponse({"results": []}, status=200)

    return JsonResponse({"results": results})

@require_GET
def ldap_search_server(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 1:
        return JsonResponse({"results": []})

    results = []
    try:
        from accounts import ldap_utils
        username = request.session.get("ldap_username")
        password = request.session.get("ldap_password")
        conn = ldap_utils._get_ldap_connection(username, password)
        base_dn = getattr(settings, "LDAP_BASE_DN", "")
        conn.search(
            search_base=base_dn,
            search_filter=f"(|(sAMAccountName=*{q}*)(cn=*{q}*)(mail=*{q}*))",
            search_scope='SUBTREE',
            attributes=['sAMAccountName', 'mail', 'cn', 'title']
        )
        for e in conn.entries:
            results.append({
                "sAMAccountName": str(getattr(e, 'sAMAccountName', '')),
                "mail": str(getattr(e, 'mail', '')),
                "cn": str(getattr(e, 'cn', '')),
                "title": str(getattr(e, 'title', '')),
            })
        try:
            conn.unbind()
        except Exception:
            pass
    except Exception as ex:
        logger.warning("LDAP search failed, falling back to users table: %s", ex)
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            like = f"%{q}%"
            cur.execute(
                "SELECT username as sAMAccountName, email as mail, username as cn "
                "FROM users WHERE username LIKE %s OR email LIKE %s LIMIT 40",
                (like, like)
            )
            rows = cur.fetchall()
            for r in rows:
                results.append({
                    "sAMAccountName": r.get("sAMAccountName"),
                    "mail": r.get("mail"),
                    "cn": r.get("cn"),
                    "title": ""
                })
        finally:
            cur.close(); conn.close()

    return JsonResponse({"results": results})

def _get_all_projects(limit=200):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name FROM projects ORDER BY created_at DESC LIMIT %s", (limit,))
        return cur.fetchall()
    finally:
        cur.close(); conn.close()

def _get_project_coe_ids(project_id):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT coe_id FROM project_coes WHERE project_id=%s", (project_id,))
        rows = cur.fetchall()
        return [r['coe_id'] for r in rows] if rows else []
    finally:
        cur.close(); conn.close()

def create_project(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        desc = (request.POST.get("description") or "").strip()
        start_date = request.POST.get("start_date") or None
        end_date = request.POST.get("end_date") or None
        pdl_username = request.POST.get("pdl_username") or None
        mapped_coe_ids = request.POST.getlist("mapped_coe_ids")

        if not name:
            users = _fetch_users()
            coes = _get_all_coes()
            projects = _get_all_projects()
            conn = get_connection()
            cur = conn.cursor(dictionary=True)
            try:
                cur.execute("SELECT id, name, coe_id FROM domains ORDER BY name")
                domains = cur.fetchall()
            finally:
                cur.close(); conn.close()
            return render(request, "projects/create_project.html", {
                "users": users, "coes": coes, "projects": projects, "domains": domains, "error": "Project name is required."
            })

        pdl_user_id = None
        if pdl_username:
            # prefer local ldap_directory email; otherwise use the supplied identifier
            pdl_user_id = None
            if pdl_username:
                local = _get_local_ldap_entry(pdl_username)
                if local:
                    pdl_user_id = local.get("email") or local.get("username")
                    try:
                        _ensure_user_from_ldap(request, pdl_user_id)
                    except Exception:
                        logger.exception("Failed to ensure users row for pdl %s", pdl_user_id)
                else:
                    pdl_user_id = pdl_username if "@" in pdl_username else pdl_username
                    try:
                        _ensure_user_from_ldap(request, pdl_username)
                    except Exception:
                        logger.exception("Failed to ensure users row for pdl (fallback) %s", pdl_username)

        conn = get_connection()
        cur = conn.cursor()
        project_id = None
        try:
            cur.execute(
                "INSERT INTO projects (name, description, start_date, end_date, pdl_user_id) VALUES (%s, %s, %s, %s, %s)",
                (name, desc or None, start_date, end_date, pdl_user_id)
            )
            conn.commit()
            project_id = cur.lastrowid
        finally:
            cur.close(); conn.close()

        try:
            int_coe_ids = [int(x) for x in mapped_coe_ids if x]
        except Exception:
            int_coe_ids = []
        if project_id and int_coe_ids:
            _replace_project_coes(project_id, int_coe_ids)

        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "project_id": project_id})
        return redirect(reverse("projects:list"))

    users = _fetch_users()
    coes = _get_all_coes()
    projects = _get_all_projects()
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name, coe_id FROM domains ORDER BY name")
        domains = cur.fetchall()
    finally:
        cur.close(); conn.close()

    return render(request, "projects/create_project.html", {
        "users": users, "coes": coes, "projects": projects, "domains": domains
    })

def edit_project(request, project_id=None):
    """
    Edit project page:
      - Shows a dropdown of projects where the logged-in user is the creator (derived from prism_wbs.creator).
      - Allows editing of fields: oem_name, pdl_user_id, pdl_name (auto), pm_user_id, pm_name (auto),
        start_date, end_date, description.
      - Uses LDAP helper get_user_entry_by_username(...) to populate CN (pdl_name/pm_name).
    """
    session_cn = request.session.get("cn", "").strip()  # e.g. "DEO Sant Anurag"
    session_ldap = request.session.get("ldap_username")
    session_pwd = request.session.get("ldap_password")
    creds = (session_ldap, session_pwd) if session_ldap and session_pwd else None

    # helper to turn "DEO Sant Anurag" -> "Sant Anurag DEO"
    def cn_to_creator(cn: str):
        if not cn:
            return ""
        parts = cn.split()
        if len(parts) >= 2:
            # last name is first token, rest are given names
            return " ".join(parts[1:]) + " " + parts[0]
        return cn

    # fetch projects where this session user is creator in prism_wbs
    editable_projects = []
    try:
        creator_name = cn_to_creator(session_cn)
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            # join prism_wbs -> projects to list unique projects where creator matches
            cur.execute("""
                SELECT DISTINCT p.id, p.name
                FROM prism_wbs pw
                JOIN projects p ON pw.project_id = p.id
                WHERE TRIM(pw.creator) = %s
                ORDER BY p.name
            """, (creator_name,))
            editable_projects = cur.fetchall() or []
        finally:
            cur.close(); conn.close()
    except Exception:
        logger.exception("Failed to fetch editable projects for creator=%s", creator_name)
        editable_projects = []

    # POST: save edits
    if request.method == "POST":
        # project_id may come from the form (dropdown)
        try:
            form_project_id = int(request.POST.get("project_choice") or project_id or 0)
        except Exception:
            return HttpResponseBadRequest("Invalid project selected")

        # ensure the selected project is in editable_projects (authorization)
        allowed_ids = {p["id"] for p in editable_projects}
        if allowed_ids and form_project_id not in allowed_ids:
            return HttpResponseForbidden("You are not authorized to edit this project")

        # gather posted values
        oem_name = (request.POST.get("oem_name") or "").strip() or None
        pdl_sel = (request.POST.get("pdl_user_id") or "").strip() or None  # we expect email primarily
        pm_sel = (request.POST.get("pm_user_id") or "").strip() or None
        start_date = request.POST.get("start_date") or None
        end_date = request.POST.get("end_date") or None
        description = (request.POST.get("description") or "").strip() or None

        # helper: ensure user exists in users table and return user_id (re-uses existing helper)
        pdl_user_id_db = None
        pm_user_id_db = None
        pdl_name_val = None
        pm_name_val = None

        # -------------------------
        # PDL handling - prefer local ldap_directory.email (store email string in projects.pdl_user_id)
        # -------------------------
        pdl_user_id_db = None   # will hold the email string (or fallback identifier)
        pdl_name_val = None
        if pdl_sel:
            # first try local ldap_directory (preferred)
            local = _get_local_ldap_entry(pdl_sel)
            if local:
                # prefer email from local directory
                pdl_user_id_db = local.get("email") or local.get("username") or pdl_sel
                pdl_name_val = local.get("cn") or local.get("username")
                # ensure users row exists (do not use its id for saving - we store email string)
                try:
                    _ensure_user_from_ldap(request, pdl_user_id_db)
                except Exception:
                    logger.exception("Failed to ensure users row for PDL %s", pdl_user_id_db)
            else:
                # fallback: if supplied value looks like an email, use it; else use supplied identifier as-is
                pdl_user_id_db = pdl_sel if "@" in pdl_sel else pdl_sel
                try:
                    _ensure_user_from_ldap(request, pdl_sel)
                except Exception:
                    logger.exception("Failed to ensure users row for PDL fallback %s", pdl_sel)

                # optional: attempt live LDAP only to fetch CN if you still want display name filled when local misses
                try:
                    if creds and creds[0] and creds[1]:
                        from accounts import ldap_utils
                        user_entry = None
                        try:
                            user_entry = get_user_entry_by_username(pdl_sel, username_password_for_conn=creds)
                        except Exception:
                            user_entry = None
                        if user_entry:
                            if hasattr(user_entry, "entry_attributes_as_dict"):
                                attrs = user_entry.entry_attributes_as_dict
                                pdl_name_val = attrs.get("cn") or attrs.get("displayName") or attrs.get("name")
                                if isinstance(pdl_name_val, (list, tuple)):
                                    pdl_name_val = pdl_name_val[0] if pdl_name_val else None
                            elif isinstance(user_entry, dict):
                                pdl_name_val = user_entry.get("cn") or user_entry.get("displayName") or user_entry.get("name")
                            else:
                                pdl_name_val = getattr(user_entry, "cn", None) or getattr(user_entry, "displayName", None)
                            if pdl_name_val:
                                pdl_name_val = str(pdl_name_val).strip()
                except Exception:
                    logger.exception("Live LDAP lookup for PDL failed for %s", pdl_sel)


        # -------------------------
        # PM handling - prefer local ldap_directory.email (store email string in projects.pm_user_id)
        # -------------------------
        pm_user_id_db = None
        pm_name_val = None
        if pm_sel:
            local = _get_local_ldap_entry(pm_sel)
            if local:
                pm_user_id_db = local.get("email") or local.get("username") or pm_sel
                pm_name_val = local.get("cn") or local.get("username")
                try:
                    _ensure_user_from_ldap(request, pm_user_id_db)
                except Exception:
                    logger.exception("Failed to ensure users row for PM %s", pm_user_id_db)
            else:
                pm_user_id_db = pm_sel if "@" in pm_sel else pm_sel
                try:
                    _ensure_user_from_ldap(request, pm_sel)
                except Exception:
                    logger.exception("Failed to ensure users row for PM fallback %s", pm_sel)

                try:
                    if creds and creds[0] and creds[1]:
                        from accounts import ldap_utils
                        user_entry = None
                        try:
                            user_entry = get_user_entry_by_username(pm_sel, username_password_for_conn=creds)
                        except Exception:
                            user_entry = None
                        if user_entry:
                            if hasattr(user_entry, "entry_attributes_as_dict"):
                                attrs = user_entry.entry_attributes_as_dict
                                pm_name_val = attrs.get("cn") or attrs.get("displayName") or attrs.get("name")
                                if isinstance(pm_name_val, (list, tuple)):
                                    pm_name_val = pm_name_val[0] if pm_name_val else None
                            elif isinstance(user_entry, dict):
                                pm_name_val = user_entry.get("cn") or user_entry.get("displayName") or user_entry.get("name")
                            else:
                                pm_name_val = getattr(user_entry, "cn", None) or getattr(user_entry, "displayName", None)
                            if pm_name_val:
                                pm_name_val = str(pm_name_val).strip()
                except Exception:
                    logger.exception("Live LDAP lookup for PM failed for %s", pm_sel)


        # persist update to projects table
        try:
            conn = get_connection()
            cur = conn.cursor()
            try:
                cur.execute("""
                    UPDATE projects
                    SET oem_name=%s,
                        pdl_user_id=%s,
                        pdl_name=%s,
                        pm_user_id=%s,
                        pm_name=%s,
                        start_date=%s,
                        end_date=%s,
                        description=%s
                    WHERE id=%s
                """, (oem_name, pdl_user_id_db, pdl_name_val, pm_user_id_db, pm_name_val, start_date, end_date, description, form_project_id))
                conn.commit()
            finally:
                cur.close(); conn.close()
            messages.success(request, "Project updated successfully.")
            # after successful save, redirect to same page to display latest details
            return redirect(reverse("projects:edit", args=[form_project_id]))
        except IntegrityError as e:
            logger.exception("Project update IntegrityError: %s", e)
            messages.error(request, "Project update failed (duplicate or constraint).")
            return redirect(reverse("projects:edit", args=[form_project_id]))
        except Exception as ex:
            logger.exception("Project update failed: %s", ex)
            messages.error(request, f"Failed to update project: {str(ex)}")
            return redirect(reverse("projects:edit", args=[form_project_id]))

    # GET: show form
    # If project_id provided, load that project's current values; else choose first editable project
    selected_project_id = project_id or (editable_projects[0]["id"] if editable_projects else None)
    project = None
    if selected_project_id:
        project = _fetch_project(selected_project_id)
    else:
        project = None

    # Also include list of editable projects for dropdown
    return render(request, "projects/edit_project.html", {
        "editable_projects": editable_projects,
        "selected_project": project,
        "ldap_username": session_ldap,
    })

@require_POST
def map_coes(request):
    """
    AJAX endpoint to map COEs to a project. Accepts:
      - project_choice: 'new' or existing project id
      - if 'new', also requires name (and optional description, start/end, pdl_username)
      - mapped_coe_ids: multiple values OK
    """
    project_choice = (request.POST.get("project_choice") or "").strip()
    selected_coes = request.POST.getlist("mapped_coe_ids")
    try:
        coe_ids = [int(x) for x in selected_coes if x]
    except Exception:
        coe_ids = []

    if project_choice == "new":
        name = (request.POST.get("name") or "").strip()
        if not name:
            return JsonResponse({"success": False, "error": "Project name required."}, status=400)
        desc = (request.POST.get("description") or "").strip()
        start_date = request.POST.get("start_date") or None
        end_date = request.POST.get("end_date") or None
        pdl_username = request.POST.get("pdl_username") or None
        pdl_user_id = None
        if pdl_username:
            pdl_user_id = _ensure_user_from_ldap(request.pdl_username)

        conn = get_connection()
        cur = conn.cursor()
        project_id = None
        try:
            cur.execute(
                "INSERT INTO projects (name, description, start_date, end_date, pdl_user_id) VALUES (%s, %s, %s, %s, %s)",
                (name, desc or None, start_date, end_date, pdl_user_id)
            )
            conn.commit()
            project_id = cur.lastrowid
        finally:
            cur.close(); conn.close()

        if project_id:
            _replace_project_coes(project_id, coe_ids)
        return JsonResponse({"success": True, "project_id": project_id})

    else:
        try:
            project_id = int(project_choice)
        except ValueError:
            return JsonResponse({"success": False, "error": "Invalid project selection."}, status=400)
        proj = _fetch_project(project_id)
        if not proj:
            return JsonResponse({"success": False, "error": "Project not found."}, status=404)
        _replace_project_coes(project_id, coe_ids)
        return JsonResponse({"success": True, "project_id": project_id})

@require_GET
def api_coes(request):
    coes = _get_all_coes()
    return JsonResponse({"coes": coes})

@require_GET
def api_projects(request):
    projects = _get_all_projects()
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT project_id, COUNT(*) as cnt FROM project_coes GROUP BY project_id")
        rows = cur.fetchall()
        counts = {r['project_id']: r['cnt'] for r in rows} if rows else {}
    finally:
        cur.close(); conn.close()
    for p in projects:
        p['mapped_coe_count'] = counts.get(p['id'], 0)
    return JsonResponse({"projects": projects})

# --- get_allocations_for_iom (returns saved rows for requested iom_id + month_start) ---
@require_GET
def get_allocations_for_iom(request):
    """
    Return saved allocations for an IOM for the canonical billing month.
    Accepts either:
      - month_start param in YYYY-MM-DD (will resolve to billing period containing that date), OR
      - month param in YYYY-MM (preferred) — will resolve billing_start via get_billing_period.
    """
    project_id = request.GET.get("project_id")
    iom_row_id = request.GET.get("iom_row_id")
    # accept month (YYYY-MM) or month_start (date string)
    month_param = request.GET.get("month")  # YYYY-MM
    month_start_param = request.GET.get("month_start")  # YYYY-MM-DD

    if not (project_id and iom_row_id):
        return JsonResponse({"ok": False, "error": "missing params"}, status=400)

    # determine canonical billing_start date to query monthly_allocation_entries
    billing_start = None
    try:
        if month_param:
            year, mon = map(int, month_param.split("-"))
            billing_start, _ = get_billing_period(year, mon)
        elif month_start_param:
            # if month_start_param corresponds to a billing window, use that window's start, else use date as-is
            try:
                dt = datetime.strptime(month_start_param, "%Y-%m-%d").date()
            except Exception:
                dt = None
            if dt:
                # find billing period containing dt (if any); fallback to dt.replace(day=1)
                try:
                    bs, be = get_billing_period_for_date(dt)
                    billing_start = bs
                except Exception:
                    billing_start = dt.replace(day=1)
        else:
            # fallback to current month's billing start
            today = date.today()
            billing_start, _ = get_billing_period(today.year, today.month)
    except Exception as exc:
        logger.exception("get_allocations_for_iom: failed to resolve billing_start: %s", exc)
        return JsonResponse({"ok": False, "error": "Invalid month parameter"}, status=400)

    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT user_ldap, total_hours
                FROM monthly_allocation_entries
                WHERE project_id=%s AND iom_id=%s AND month_start=%s
            """, (project_id, iom_row_id, billing_start))
            rows = cur.fetchall() or []
    except Exception as exc:
        logger.exception("get_allocations_for_iom failed: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)

    saved_items = [{"user_ldap": r[0], "total_hours": float(r[1] or 0)} for r in rows]
    return JsonResponse({"ok": True, "saved_items": saved_items, "billing_start": billing_start.strftime("%Y-%m-%d")})


@require_POST
def save_monthly_allocations(request):
    """
    Save or update monthly allocation entries, using billing cycle start_date as canonical month_start.

    Accepts JSON payload (preferred) or standard form POST.

    JSON examples (both supported):
      { "project_id": 46, "month": "2025-08", "items": [ {"iom_id":"PEU_...", "user_ldap":"x@y", "total_hours":183.75}, ... ] }

      or

      { "project_id": 46, "month_start": "2025-07-21", "items": [ ... ] }

    Legacy/form support:
      - form fields project_id, month (YYYY-MM) OR year+month_num
      - or a JSON string in items_json field
      - or patterned POST fields iom_id1,user_ldap1,hours1, etc.
    """
    try:
        # Read body and attempt JSON
        body_raw = request.body.decode("utf-8").strip()
        data = {}
        if body_raw:
            try:
                data = json.loads(body_raw)
            except Exception:
                data = {}

        # Primary inputs
        project_id = data.get("project_id") or request.POST.get("project_id") or request.GET.get("project_id")
        month_param = data.get("month") or request.POST.get("month") or request.GET.get("month")  # expected "YYYY-MM"
        month_start_param = data.get("month_start") or request.POST.get("month_start") or request.GET.get("month_start")  # expected "YYYY-MM-DD"

        # Allow year + month numeric fallback
        if not month_param:
            year_f = request.POST.get("year") or request.GET.get("year") or data.get("year")
            mon_f = request.POST.get("month_num") or request.POST.get("mon") or request.GET.get("month_num") or data.get("month_num")
            if year_f and mon_f:
                try:
                    month_param = f"{int(year_f):04d}-{int(mon_f):02d}"
                except Exception:
                    month_param = None

        # If still missing, see if items_json contains month/project info
        if (not project_id or (not month_param and not month_start_param)) and (data.get("items") or data.get("items_json") or request.POST.get("items_json")):
            try:
                items_blob = data.get("items") or data.get("items_json") or request.POST.get("items_json")
                if isinstance(items_blob, str):
                    parsed_blob = json.loads(items_blob)
                else:
                    parsed_blob = items_blob
                if isinstance(parsed_blob, dict):
                    project_id = project_id or parsed_blob.get("project_id")
                    month_param = month_param or parsed_blob.get("month")
                    month_start_param = month_start_param or parsed_blob.get("month_start")
            except Exception:
                pass

        # minimal validation
        if not project_id or (not month_param and not month_start_param):
            logger.warning("save_monthly_allocations missing project or month; data_keys=%s POST_keys=%s body_len=%d",
                           list(data.keys()) if isinstance(data, dict) else None,
                           list(request.POST.keys()), len(body_raw))
            return JsonResponse({"ok": False, "error": "Missing project_id or month/month_start"}, status=400)

        # coerce project_id
        try:
            project_id = int(project_id)
        except Exception:
            return JsonResponse({"ok": False, "error": "Invalid project_id"}, status=400)

        # Resolve canonical billing_start & billing_end (single source of truth)
        billing_start = None
        billing_end = None
        if month_start_param:
            # parse provided date and then map to billing period that contains that date
            try:
                dt = datetime.strptime(str(month_start_param), "%Y-%m-%d").date()
                bs, be = get_billing_period_for_date(dt)
                billing_start, billing_end = bs, be
            except Exception:
                # fallback: try parse ignoring time portion
                try:
                    # attempt to parse more liberal formats
                    dt_try = datetime.strptime(str(month_start_param).split(" ")[0], "%Y-%m-%d").date()
                    bs, be = get_billing_period_for_date(dt_try)
                    billing_start, billing_end = bs, be
                except Exception:
                    return JsonResponse({"ok": False, "error": "Invalid month_start format (expected YYYY-MM-DD)"}, status=400)
        else:
            # month_param expected "YYYY-MM"
            try:
                year, mon = map(int, str(month_param).split("-"))
                billing_start, billing_end = get_billing_period(year, mon)
            except Exception:
                return JsonResponse({"ok": False, "error": "Invalid month format (expected YYYY-MM)"}, status=400)

        # Build items list
        items = []
        if isinstance(data.get("items"), list):
            items = data.get("items")
        else:
            items_json_field = data.get("items_json") or request.POST.get("items_json")
            if items_json_field:
                try:
                    if isinstance(items_json_field, str):
                        items = json.loads(items_json_field)
                    else:
                        items = items_json_field
                except Exception:
                    items = []

            # patterned form extraction (iom_id1,user_ldap1,hours1)
            if not items:
                # find keys that look like iom_id{n}
                iom_keys = [k for k in request.POST.keys() if k.startswith("iom_id")]
                if iom_keys:
                    suffixes = set()
                    for k in iom_keys:
                        s = k.replace("iom_id", "").strip("_")
                        suffixes.add(s)
                    for s in suffixes:
                        iom = request.POST.get(f"iom_id{s}")
                        user_ldap = request.POST.get(f"user_ldap{s}") or request.POST.get(f"resource{s}")
                        hours = request.POST.get(f"hours{s}") or request.POST.get(f"total_hours{s}")
                        if iom and user_ldap:
                            try:
                                hrs = float(hours) if hours not in (None, "") else 0.0
                            except Exception:
                                hrs = 0.0
                            items.append({"iom_id": iom, "user_ldap": user_ldap, "total_hours": hrs})

        if not items:
            return JsonResponse({"ok": False, "error": "No allocation items provided"}, status=400)

        # Normalize & group by iom_id
        grouped = {}
        for it in items:
            iom_id = (it.get("iom_id") or it.get("iom") or "").strip()
            user_ldap = (it.get("user_ldap") or it.get("user") or it.get("resource") or "").strip()
            if not iom_id or not user_ldap:
                continue
            try:
                hrs_raw = it.get("total_hours", it.get("hours", 0))
                hrs = float(hrs_raw) if hrs_raw not in (None, "") else 0.0
            except Exception:
                hrs = 0.0
            grouped.setdefault(iom_id, []).append((user_ldap, round(float(hrs), 2)))

        if not grouped:
            return JsonResponse({"ok": False, "error": "No valid allocation rows found"}, status=400)

        saved = {}
        with transaction.atomic():
            with connection.cursor() as cur:
                for iom_id, rows in grouped.items():
                    # Delete previous entries for this project/iom and canonical billing_start
                    cur.execute("""
                        DELETE FROM monthly_allocation_entries
                        WHERE project_id = %s AND iom_id = %s AND month_start = %s
                    """, [project_id, iom_id, billing_start])

                    # Insert new entries using canonical billing_start
                    for ldap, hrs in rows:
                        cur.execute("""
                            INSERT INTO monthly_allocation_entries
                            (project_id, iom_id, month_start, user_ldap, total_hours, created_at)
                            VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                        """, [project_id, iom_id, billing_start, ldap, str(float(hrs))])

                    # read back saved rows for reporting
                    cur.execute("""
                        SELECT user_ldap, total_hours
                        FROM monthly_allocation_entries
                        WHERE project_id = %s AND iom_id = %s AND month_start = %s
                    """, [project_id, iom_id, billing_start])
                    fetched = cur.fetchall()
                    saved[iom_id] = [{"user_ldap": r[0], "total_hours": float(r[1] or 0)} for r in fetched]

        # Return canonical billing_start/billing_end so UI knows what server used
        return JsonResponse({
            "ok": True,
            "billing_start": billing_start.strftime("%Y-%m-%d"),
            "billing_end": billing_end.strftime("%Y-%m-%d"),
            "saved": saved
        })

    except Exception as exc:
        logger.exception("save_monthly_allocations failed: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


# ---- Helper utilities ---------------------------------------------------

def _sql_in_clause(items):
    """
    Return (sql_fragment, params_list) for an IN clause for psycopg/MySQL paramstyle (%s).
    If items is empty returns ("(NULL)", []) to produce a false IN clause safely.
    """
    if not items:
        return "(NULL)", []
    placeholders = ",".join(["%s"] * len(items))
    return f"({placeholders})", list(items)


def is_pdl_user(ldap_entry):
    """
    Determine whether the LDAP user is a PDL (Project Delivery Lead) or manager.
    This is a conservative check and should be replaced/extended based on your LDAP schema:
      - check memberOf for specific group
      - check 'title', 'employeeType', or a custom attr like 'role'
    ldap_entry is expected to be the object returned by get_user_entry_by_username.
    """
    if not ldap_entry:
        return False

    # try a few common attributes (adjust to your environment)
    try:
        # If your LDAP helper returns a dict-like or attribute accessor, adapt accordingly
        attrs = {}
        if hasattr(ldap_entry, "entry_attributes_as_dict"):
            attrs = ldap_entry.entry_attributes_as_dict
        elif isinstance(ldap_entry, dict):
            attrs = ldap_entry
        else:
            # fallback: try to access attribute names directly
            # create attrs by reading typical attr names if present
            for name in ("title", "employeeType", "memberOf", "role"):
                val = getattr(ldap_entry, name, None)
                if val:
                    attrs[name] = val

        # If explicit role attribute mentions PDL/Manager
        role_val = (attrs.get("employeeType") or attrs.get("title") or attrs.get("role") or "")
        if isinstance(role_val, (list, tuple)):
            role_val = " ".join(role_val)
        if role_val and ("pdl" in role_val.lower() or "project delivery" in role_val.lower() or "manager" in role_val.lower()):
            return True

        # If memberOf contains a PDL/Managers group
        member_of = attrs.get("memberOf") or attrs.get("memberof") or []
        if isinstance(member_of, str):
            member_of = [member_of]
        for grp in member_of:
            if "pdl" in grp.lower() or "manager" in grp.lower() or "project-delivery" in grp.lower():
                return True
    except Exception:
        logger.exception("is_pdl_user: unexpected structure for ldap_entry")

    return False


# ---- Main view ---------------------------------------------------------

def team_allocations(request):
    """
    Team Allocation page (billing-period aware). Uses get_billing_period(year, month)
    when `month` query param provided (YYYY-MM). If no monthly_hours_limit entry exists
    it falls back to calendar month. Builds rows from monthly_allocation_entries where
    mae.month_start = billing_start (canonical).
    """
    # SINGLE source of truth for LDAP username/password from session
    session_ldap = request.session.get("ldap_username")
    session_pwd = request.session.get("ldap_password")
    logger.debug("team_allocations - session_ldap: %s", session_ldap)

    # require login
    if not session_ldap or not session_pwd:
        return redirect("accounts:login")
    creds = (session_ldap, session_pwd)

    # --- month param parsing (use YYYY-MM) ---
    month_str = request.GET.get("month")
    if month_str:
        try:
            year, mon = map(int, month_str.split("-"))
            month_start, month_end = get_billing_period(year, mon)
        except Exception:
            logger.exception("team_allocations: invalid month param '%s'", month_str)
            today = date.today()
            month_start, month_end = get_billing_period(today.year, today.month)

    else:
        # default - current month billing period fallback
        today = date.today()
        month_start, month_end = get_billing_period(today.year, today.month)

    # --- get LDAP user entry -----------------------------------------------------
    user_entry = get_user_entry_by_username(session_ldap, username_password_for_conn=creds)
    if not user_entry:
        logger.warning("team_allocations: user_entry not found for %s", session_ldap)
        return redirect("accounts:login")

    # --- get reportees via LDAP --------------------------------------------------
    reportees_entries = get_reportees_for_user_dn(getattr(user_entry, "entry_dn", None),
                                                 username_password_for_conn=creds) or []
    reportees_ldaps = []
    for ent in reportees_entries:
        val = None
        if isinstance(ent, dict):
            val = ent.get("userPrincipalName") or ent.get("mail") or ent.get("userid") or ent.get("sAMAccountName")
        else:
            for attr in ("userPrincipalName", "mail", "sAMAccountName", "uid"):
                val = getattr(ent, attr, None) or val
        if val:
            reportees_ldaps.append(str(val).strip())

    # include manager themselves if PDL
    try:
        if is_pdl_user(user_entry):
            if session_ldap not in reportees_ldaps:
                reportees_ldaps.append(session_ldap)
                logger.debug("team_allocations: user is PDL, added own ldap to reportees list")
    except Exception:
        logger.exception("team_allocations: error checking PDL role for %s", session_ldap)

    rows = []
    if reportees_ldaps:
        in_clause, in_params = _sql_in_clause(reportees_ldaps)
        sql = f"""
            SELECT mae.id AS item_id,
                   mae.id AS allocation_id,
                   mae.user_ldap,
                   u.username, u.email,
                   p.name AS project_name,
                   pw.department AS domain_name,
                   COALESCE(mae.total_hours, 0.00) AS total_hours
            FROM monthly_allocation_entries mae
            LEFT JOIN users u ON u.email = mae.user_ldap
            LEFT JOIN projects p ON mae.project_id = p.id
            LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
            WHERE mae.month_start = %s
              AND mae.user_ldap IN {in_clause}
            ORDER BY u.username, p.name
        """
        params = [month_start] + in_params
        try:
            with connection.cursor() as cur:
                cur.execute(sql, params)
                rows = dictfetchall(cur) or []
                logger.debug("team_allocations: fetched %d allocation rows", len(rows))
        except Exception as exc:
            logger.exception("team_allocations: DB query failed: %s", exc)
            rows = []
    else:
        logger.debug("team_allocations: no reportees to fetch for %s", session_ldap)

    # deduplicate and attach weekly allocations as before (existing code)
    dedup = {}
    for r in rows:
        key = r.get("item_id") or (r.get("allocation_id"), r.get("user_ldap"), r.get("project_name"))
        if key not in dedup:
            dedup[key] = r
    all_rows = list(dedup.values())

    allocation_ids = list({r["allocation_id"] for r in all_rows if r.get("allocation_id")})
    weekly_map = {}
    if allocation_ids:
        in_clause, in_params = _sql_in_clause(allocation_ids)
        sql = f"""
            SELECT allocation_id, week_number, percent, hours, status
            FROM weekly_allocations
            WHERE allocation_id IN {in_clause}
        """
        try:
            with connection.cursor() as cur:
                cur.execute(sql, in_params)
                for r in dictfetchall(cur) or []:
                    try:
                        alloc = r.get("allocation_id")
                        wk = int(r.get("week_number") or 0)
                        pct = float(r.get("percent") or 0.0)
                        hrs_raw = r.get("hours") or 0.0
                        try:
                            hrs = float(hrs_raw)
                        except Exception:
                            hrs = 0.0
                        status = r.get("status") or ""
                        if alloc is None or wk <= 0:
                            continue
                        weekly_map.setdefault(alloc, {})[wk] = {"percent": pct, "hours": hrs, "status": status}
                    except Exception:
                        logger.exception("team_allocations: bad weekly row %r", r)
        except Exception as exc:
            logger.exception("team_allocations: weekly allocations query failed: %s", exc)
            weekly_map = {}

    # attach weekly attrs and compute display_name
    for r in all_rows:
        aid = r.get("allocation_id")
        wk = weekly_map.get(aid, {})
        r["w1"] = wk.get(1, {}).get("percent", 0)
        r["w2"] = wk.get(2, {}).get("percent", 0)
        r["w3"] = wk.get(3, {}).get("percent", 0)
        r["w4"] = wk.get(4, {}).get("percent", 0)
        r["s1"] = wk.get(1, {}).get("status", "")
        r["s2"] = wk.get(2, {}).get("status", "")
        r["s3"] = wk.get(3, {}).get("status", "")
        r["s4"] = wk.get(4, {}).get("status", "")

        display = (r.get("username") or r.get("user_ldap") or "")
        if r.get("email"):
            display += f" <{r['email']}>"
        r["display_name"] = display

    # Determine reportees with no allocations
    allocated_ldaps = {(r.get("user_ldap") or "").strip().lower() for r in all_rows if (r.get("user_ldap") or "").strip()}
    reportees_no_alloc = []
    try:
        for ld in reportees_ldaps:
            key = (ld or "").strip()
            if not key:
                continue
            if key.lower() not in allocated_ldaps:
                local = _get_local_ldap_entry(key)
                display_name = local.get("cn") if local and local.get("cn") else (local.get("username") if local else "")
                email_val = local.get("email") if local and local.get("email") else (key if "@" in key else "")
                reportees_no_alloc.append({
                    "ldap": key,
                    "display": display_name,
                    "email": email_val
                })
    except Exception:
        logger.exception("team_allocations: error building reportees_no_alloc")
        reportees_no_alloc = []

    return render(request, "projects/team_allocations.html", {
        "rows": all_rows,
        "weekly_map": weekly_map,
        "month_start": month_start,
        "reportees": reportees_ldaps,
        "reportees_no_alloc": reportees_no_alloc,
    })



# -------------------------
# save_team_allocation
# -------------------------
@require_POST
def save_team_allocation(request):
    """
    Save weekly percent allocations for a monthly allocation (monthly_allocation_entries.id).
    Expects JSON body: { "allocation_id": 123, "weekly": { "1": 25.0, "2": 25.0, "3": 25.0, "4": 25.0 } }
    Returns JSON: { ok: True, allocation_id: 123, weeks: { "1": "46.25", ... } }
    """
    # parse JSON payload
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON payload")

    allocation_id = payload.get('allocation_id')
    weekly = payload.get('weekly', {}) or {}

    try:
        allocation_id = int(allocation_id)
    except Exception:
        return HttpResponseBadRequest("Invalid allocation_id")

    if allocation_id <= 0:
        return HttpResponseBadRequest("Invalid allocation_id")

    # fetch canonical allocation info from monthly_allocation_entries
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id, total_hours, user_ldap
            FROM monthly_allocation_entries
            WHERE id = %s
            LIMIT 1
        """, [allocation_id])
        rec = cur.fetchone()

    if not rec:
        return HttpResponseBadRequest("Allocation not found")

    _, total_hours_raw, user_ldap = rec

    # coerce to Decimal for accurate arithmetic
    try:
        total_hours_dec = Decimal(str(total_hours_raw or '0.00'))
    except Exception:
        total_hours_dec = Decimal('0.00')

    result_weeks = {}

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                for wk_key, pct_val in weekly.items():
                    # normalize week number
                    try:
                        week_num = int(wk_key)
                    except Exception:
                        # skip invalid week keys
                        continue
                    # coerce percent to Decimal and clamp
                    try:
                        pct_dec = Decimal(str(pct_val))
                    except Exception:
                        pct_dec = Decimal('0.00')
                    if pct_dec < Decimal('0.00'):
                        pct_dec = Decimal('0.00')
                    if pct_dec > Decimal('100.00'):
                        pct_dec = Decimal('100.00')

                    # compute hours = total_hours * (pct/100), quantized to 2 decimals
                    hours_dec = (total_hours_dec * (pct_dec / Decimal('100.00'))).quantize(
                        Decimal('0.01'), rounding=ROUND_HALF_UP
                    )

                    # Upsert percent and hours
                    cur.execute("""
                        INSERT INTO weekly_allocations (allocation_id, week_number, percent, hours, updated_at)
                        VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                        ON DUPLICATE KEY UPDATE
                          percent = VALUES(percent),
                          hours = VALUES(hours),
                          updated_at = CURRENT_TIMESTAMP
                    """, [allocation_id, week_num, str(pct_dec), str(hours_dec)])

                    # prepare response payload (strings to preserve decimal formatting)
                    result_weeks[str(week_num)] = format(hours_dec, '0.2f')

    except Exception as exc:
        # optional: logger.exception("save_team_allocation failed: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)})

    return JsonResponse({"ok": True, "allocation_id": allocation_id, "weeks": result_weeks})

# -------------------------------------------------------------------
# 3. MY ALLOCATIONS (VIEW)
# -------------------------------------------------------------------
def my_allocations(request):
    """
    Billing-cycle-aware my_allocations view that always builds week blocks (1..4)
    for the billing period (start_date..end_date). If weekly_allocations rows are missing
    we provide a fallback equal-split of total_hours so daily punching and Save Week
    buttons remain available.
    """
    from decimal import Decimal

    # Resolve user identity (same approach you used previously)
    session_ldap = request.session.get("ldap_username") or request.session.get("user_ldap") or getattr(request.user, "email", None)
    if not session_ldap:
        return HttpResponseBadRequest("No user identity found")

    # month param: YYYY-MM
    month_param = request.GET.get("month", date.today().strftime("%Y-%m"))
    try:
        year, month = map(int, month_param.split("-"))
    except Exception:
        year, month = date.today().year, date.today().month

    # Get canonical billing period
    billing_start, billing_end = get_billing_period(year, month)

    # Fetch allocations for this user for the canonical billing_start
    with connection.cursor() as cur:
        cur.execute("""
            SELECT mae.id AS allocation_id,
                   mae.project_id,
                   p.name AS project_name,
                   mae.iom_id,
                   pw.department AS domain_name,
                   COALESCE(mae.total_hours, 0.00) AS total_hours
            FROM monthly_allocation_entries mae
            LEFT JOIN projects p ON mae.project_id = p.id
            LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
            WHERE mae.user_ldap = %s AND mae.month_start = %s
            ORDER BY p.name
        """, [session_ldap, billing_start])
        alloc_rows = dictfetchall(cur)

    allocation_ids = [r['allocation_id'] for r in alloc_rows] if alloc_rows else []

    # Fetch weekly_allocations rows for allocation_ids (if any)
    weekly_alloc = {}  # map allocation_id -> week_number -> hours
    if allocation_ids:
        in_clause = ",".join(["%s"] * len(allocation_ids))
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT allocation_id, week_number, COALESCE(hours,0) as hours
                FROM weekly_allocations
                WHERE allocation_id IN ({in_clause})
            """, allocation_ids)
            for r in dictfetchall(cur):
                aid = r['allocation_id']
                weekly_alloc.setdefault(aid, {})[int(r['week_number'])] = Decimal(str(r['hours'] or '0.00'))

    # Fetch user punches in the billing window for these allocations
    user_punch_map_daily = {}  # allocation_id -> iso_date -> Decimal(hours)
    if allocation_ids:
        in_clause = ",".join(["%s"] * len(allocation_ids))
        params = [session_ldap] + allocation_ids + [billing_start, billing_end]
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT allocation_id, punch_date, actual_hours
                FROM user_punches
                WHERE user_ldap = %s
                  AND allocation_id IN ({in_clause})
                  AND punch_date BETWEEN %s AND %s
            """, params)
            for r in dictfetchall(cur):
                aid = r['allocation_id']
                d = r['punch_date']
                iso = d.strftime("%Y-%m-%d")
                user_punch_map_daily.setdefault(aid, {})[iso] = Decimal(str(r['actual_hours'] or '0.00'))

    # Build daily_dates list for billing period and compute week_number relative to billing_start
    daily_dates = []
    cur_day = billing_start
    # compute total weeks in billing period (dynamic)
    total_days = (billing_end - billing_start).days + 1
    total_weeks = int(ceil(total_days / 7.0))
    while cur_day <= billing_end:
        week_number = month_day_to_week_number_for_period(cur_day, billing_start, billing_end)
        daily_dates.append({
            "date": cur_day,
            "iso": cur_day.strftime("%Y-%m-%d"),
            "week_number": week_number,
            "is_weekend": cur_day.weekday() >= 5,
        })
        cur_day += timedelta(days=1)

    # For each allocation build the row structure consumed by template
    rows = []
    for r in alloc_rows:
        aid = r['allocation_id']
        total_hours = Decimal(str(r.get('total_hours') or '0.00'))

        # Determine week allocation hours: prefer DB weekly_alloc rows; if missing, fallback to equal split
        weeks = {}
        db_weeks = weekly_alloc.get(aid, {})
        if db_weeks:
            # Use DB values; ensure all 1..4 keys exist (0 if missing)
            for wk in range(1,5):
                weeks[wk] = db_weeks.get(wk, Decimal('0.00'))
        else:
            # Fallback: split total_hours equally across 4 weeks (rounding to 2 decimals)
            if total_hours > 0:
                per_week = (total_hours / Decimal(4)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                # Correct rounding so sum equals total_hours (adjust last week)
                weeks = {wk: per_week for wk in range(1,5)}
                sum_weeks = sum(weeks.values())
                diff = total_hours - sum_weeks
                if diff != Decimal('0.00'):
                    weeks[4] = (weeks[4] + diff).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            else:
                weeks = {wk: Decimal('0.00') for wk in range(1,5)}

        # Compute which weeks to show as 'present' (if week alloc >0 or total_hours >0 show them)
        weeks_present = [wk for wk,h in weeks.items() if h > 0] or [1,2,3,4]

        # Compute punched per week (sum of daily punches within week range)
        punched_per_week = {}
        for wk in range(1,5):
            # week start and end relative to billing_start
            wk_start = billing_start + timedelta(days=(wk-1)*7)
            wk_end = min(wk_start + timedelta(days=6), billing_end)
            s = Decimal('0.00')
            # sum daily punches in the map for this allocation
            for iso, hrs in (user_punch_map_daily.get(aid) or {}).items():
                d = datetime.strptime(iso, "%Y-%m-%d").date()
                if wk_start <= d <= wk_end:
                    s += Decimal(str(hrs or '0.00'))
            punched_per_week[wk] = s

        # Prepare final row (hours as strings for template)
        row = {
            "allocation_id": aid,
            "project_name": r.get('project_name'),
            "domain_name": r.get('domain_name'),
            "total_hours": format(total_hours, '0.2f'),
            "weeks_present": weeks_present,
            "w1_alloc_hours": format(weeks[1], '0.2f'),
            "w2_alloc_hours": format(weeks[2], '0.2f'),
            "w3_alloc_hours": format(weeks[3], '0.2f'),
            "w4_alloc_hours": format(weeks[4], '0.2f'),
            "w1_punched_hours": format(punched_per_week.get(1, Decimal('0.00')), '0.2f'),
            "w2_punched_hours": format(punched_per_week.get(2, Decimal('0.00')), '0.2f'),
            "w3_punched_hours": format(punched_per_week.get(3, Decimal('0.00')), '0.2f'),
            "w4_punched_hours": format(punched_per_week.get(4, Decimal('0.00')), '0.2f'),
            # wbs options used by template (attempt to load seller/buyer for iom_id)
            "wbs_options": []
        }

        # load wbs options (seller/buyer) if iom_id available
        iom_id = r.get('iom_id')
        if iom_id:
            with connection.cursor() as cur_w:
                cur_w.execute("SELECT seller_wbs_cc, buyer_wbs_cc FROM prism_wbs WHERE iom_id=%s LIMIT 1", [iom_id])
                wrow = cur_w.fetchone()
                if wrow:
                    if wrow[0]:
                        row['wbs_options'].append({"code": f"seller:{wrow[0]}", "label": f"Seller WBS: {wrow[0]}"})
                    if wrow[1]:
                        row['wbs_options'].append({"code": f"buyer:{wrow[1]}", "label": f"Buyer WBS: {wrow[1]}"})

        rows.append(row)

    # daily_map for template (string formatted)
    daily_map = {}
    for r in alloc_rows:
        aid = r['allocation_id']
        daymap = {}
        for d in daily_dates:
            iso = d['iso']
            val = user_punch_map_daily.get(aid, {}).get(iso, Decimal('0.00'))
            daymap[iso] = format(Decimal(val), '0.2f')
        daily_map[aid] = daymap

    # holidays map between billing_start and billing_end if you store holidays
    with connection.cursor() as cur:
        cur.execute("SELECT holiday_date, name FROM holidays WHERE holiday_date BETWEEN %s AND %s", [billing_start, billing_end])
        holiday_rows = dictfetchall(cur)
    holidays_map = {r['holiday_date'].strftime("%Y-%m-%d"): r['name'] for r in holiday_rows}

    return render(request, "projects/my_allocations.html", {
        "rows": rows,
        "daily_dates": daily_dates,
        "daily_map": daily_map,
        "month_start": billing_start,
        "holidays_map": holidays_map,
    })




# ---------- save weekly punches endpoint ----------
@require_POST
def save_my_alloc_weekly(request):
    """
    Expects JSON: { allocation_id: int, week_number:int, actual_hours: number, wbs: optional }
    It will upsert (INSERT .. ON DUPLICATE KEY UPDATE) into weekly_allocations table.
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        allocation_id = int(payload.get("allocation_id", 0))
        week_number = int(payload.get("week_number", 0))
        hours = Decimal(str(payload.get("actual_hours", "0"))).quantize(Decimal("0.01"), ROUND_HALF_UP)
        wbs = payload.get("wbs")
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid payload"}, status=400)

    if allocation_id <= 0 or week_number not in (1,2,3,4):
        return JsonResponse({"ok": False, "error": "Invalid allocation_id or week_number"}, status=400)

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # Upsert pattern for weekly_allocations (assuming unique key on allocation_id+week_number)
                cur.execute("""
                    INSERT INTO weekly_allocations (allocation_id, week_number, percent, hours, created_at)
                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                    ON DUPLICATE KEY UPDATE hours = VALUES(hours)
                """, [allocation_id, week_number, 0.0, str(hours)])
        return JsonResponse({"ok": True, "allocation_id": allocation_id, "week_number": week_number, "hours": f"{hours:.2f}"})
    except Exception as e:
        logger.exception("save_my_alloc_weekly failed: %s", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

# save_daily endpoint (modified to use billing period lookup for punch_date)
# -------------------------
@require_POST
def save_my_alloc_daily(request):
    """Save daily punches aligned to billing cycle."""
    try:
        data = json.loads(request.body.decode("utf-8"))
        allocation_id = int(data.get("allocation_id"))
        punch_date = datetime.strptime(data.get("punch_date"), "%Y-%m-%d").date()
        actual_hours = Decimal(str(data.get("actual_hours", 0))).quantize(Decimal("0.01"), ROUND_HALF_UP)
        wbs = data.get("wbs")

        user_ldap = request.session.get("ldap_username")

        billing_start, billing_end = get_billing_period_for_date(punch_date)
        week_number = ((punch_date - billing_start).days // 7) + 1

        # get weekly allocation
        with connection.cursor() as cur:
            cur.execute("""
                SELECT hours FROM weekly_allocations
                WHERE allocation_id=%s AND week_number=%s
            """, [allocation_id, week_number])
            rec = cur.fetchone()
        if not rec:
            return JsonResponse({"ok": False, "error": "No weekly allocation found"}, status=400)

        alloc_hours = Decimal(str(rec[0] or "0.00"))
        wk_start = billing_start + timedelta(days=(week_number - 1) * 7)
        wk_end = min(wk_start + timedelta(days=6), billing_end)

        # validate total within week
        with connection.cursor() as cur:
            cur.execute("""
                SELECT COALESCE(SUM(actual_hours),0)
                FROM user_punches
                WHERE user_ldap=%s AND allocation_id=%s AND punch_date BETWEEN %s AND %s
            """, [user_ldap, allocation_id, wk_start, wk_end])
            sum_existing = Decimal(str(cur.fetchone()[0] or "0.00"))

            cur.execute("""
                SELECT actual_hours FROM user_punches
                WHERE user_ldap=%s AND allocation_id=%s AND punch_date=%s
            """, [user_ldap, allocation_id, punch_date])
            existing_same = cur.fetchone()
            existing_same_val = Decimal(str(existing_same[0])) if existing_same else Decimal("0.00")

        total_after = (sum_existing - existing_same_val) + actual_hours
        if total_after > alloc_hours:
            return JsonResponse({"ok": False, "error": f"Exceeds weekly allocation {alloc_hours:.2f}"}, status=400)

        with transaction.atomic():
            with connection.cursor() as cur:
                cur.execute("""
                    INSERT INTO user_punches
                    (user_ldap, allocation_id, punch_date, week_number, actual_hours, wbs, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,CURRENT_TIMESTAMP)
                    ON DUPLICATE KEY UPDATE
                      actual_hours=VALUES(actual_hours),
                      wbs=VALUES(wbs),
                      updated_at=CURRENT_TIMESTAMP
                """, [user_ldap, allocation_id, punch_date, week_number, str(actual_hours), wbs])

        return JsonResponse({"ok": True, "allocation_id": allocation_id})

    except Exception as e:
        logger.exception("save_my_alloc_daily failed: %s", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)
# my_allocations_update_status
# -------------------------
@require_POST
def my_allocations_update_status(request):
    """
    Update status (ACCEPTED/REJECTED) for weeks for the logged-in user's allocation.
    """
    session_ldap = request.session.get("ldap_username")
    print("my_allocations_update_status - session_ldap:", session_ldap)
    if not session_ldap:
        return HttpResponseForbidden("Missing LDAP session username")

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    allocation_id = payload.get("allocation_id")
    updates = payload.get("updates", {})

    if not allocation_id or not isinstance(updates, dict):
        return HttpResponseBadRequest("allocation_id and updates required")

    # verify allocation belongs to logged in user
    with connection.cursor() as cur:
        cur.execute("""
            SELECT a.id, ai.user_ldap, a.total_hours
            FROM allocations a
            JOIN allocation_items ai ON ai.allocation_id = a.id
            WHERE a.id = %s LIMIT 1
        """, [allocation_id])
        rec = cur.fetchone()
        if not rec:
            return HttpResponseBadRequest("Invalid allocation_id")
        db_alloc_id, db_user_ldap, total_hours = rec

    if (db_user_ldap or "").strip() != (session_ldap or "").strip():
        return HttpResponseForbidden("You are not authorized to update this allocation")

    try:
        with transaction.atomic():
            for week_str, action in updates.items():
                try:
                    week_num = int(week_str)
                except Exception:
                    continue
                act = (action or "").strip().upper()
                if act not in ("ACCEPT", "ACCEPTED", "REJECT", "REJECTED"):
                    continue
                status_val = "ACCEPTED" if act.startswith("ACCE") else "REJECTED"
                with connection.cursor() as cur:
                    cur.execute("SELECT hours FROM weekly_allocations WHERE allocation_id = %s AND week_number = %s LIMIT 1",
                                [allocation_id, week_num])
                    hh = cur.fetchone()
                    hours_val = int(hh[0]) if hh and hh[0] is not None else 0
                    cur.execute("""
                        INSERT INTO weekly_allocations (allocation_id, week_number, hours, status)
                        VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE hours = VALUES(hours), status = VALUES(status), updated_at = CURRENT_TIMESTAMP
                    """, [allocation_id, week_num, hours_val, status_val])
        return JsonResponse({"ok": True})
    except Exception as exc:
        logger.exception("my_allocations_update_status failed: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)



# Replace existing _get_user_projects_for_allocations and monthly_allocations with these:

def _cn_to_creator(cn: str):
    """
    Convert "LASTNAME FirstName ..." -> "FirstName ... LASTNAME"
    (same conversion used in edit_project() so creator matching is consistent)
    """
    if not cn:
        return ""
    parts = str(cn).strip().split()
    if len(parts) >= 2:
        # put the first token (last name) at the end
        return " ".join(parts[1:]) + " " + parts[0]
    return str(cn).strip()


def _get_user_projects_for_allocations(request):
    """
    Return list of projects (dicts with id,name) where session user is:
      - PDL (projects.pdl_user_id == session ldap identifier string)
      - OR creator of any prism_wbs rows (prism_wbs.creator matches converted CN or ldap username)
    This aligns with project_list / edit_project logic (pdl_user_id stored as email/username string).
    """
    session_ldap = request.session.get("ldap_username")
    session_cn = request.session.get("cn", "")
    creator_name = _cn_to_creator(session_cn) if session_cn else ""

    sql = """
        SELECT DISTINCT p.id, p.name
        FROM projects p
        LEFT JOIN prism_wbs pw ON pw.project_id = p.id
        WHERE 1=0
    """
    params = []

    # match pdl_user_id string directly (projects.pdl_user_id is stored as email/username)
    if session_ldap:
        sql += " OR p.pdl_user_id = %s"
        params.append(session_ldap)

    # match prism_wbs.creator converted CN (or other creators if you need to add more)
    if creator_name:
        sql += " OR TRIM(pw.creator) = %s"
        params.append(creator_name)

    sql += " ORDER BY p.name"

    try:
        with connection.cursor() as cur:
            cur.execute(sql, params)
            projects = dictfetchall(cur)
    except Exception:
        logger.exception("Error in _get_user_projects_for_allocations")
        projects = []

    # normalize to simple dicts for template (id,name)
    out = [{"id": p.get("id"), "name": p.get("name") or ""} for p in projects]
    return out


def monthly_allocations(request):
    """
    Render monthly allocations page (billing-period aware). Accepts ?month=YYYY-MM
    and uses billing_start resolved via get_billing_period(year, month).
    """
    logger.debug("monthly_allocations called")
    session_ldap = request.session.get("ldap_username")

    # Parse month param (YYYY-MM) -> billing_start
    month_str = request.GET.get("month")
    if month_str:
        try:
            year, mon = map(int, month_str.split("-"))
            month_start, month_end = get_billing_period(year, mon)
        except Exception:
            logger.exception("monthly_allocations: invalid month param '%s'", month_str)
            from datetime import date
            today = date.today();
            month_start, month_end = get_billing_period(today.year, today.month)
    else:
        from datetime import date
        today = date.today()
        month_start, month_end = get_billing_period(today.year, today.month)

    project_id_param = request.GET.get("project_id")
    try:
        active_project_id = int(project_id_param) if project_id_param else 0
    except Exception:
        active_project_id = 0

    # Fetch projects user can allocate for
    projects = _get_user_projects_for_allocations(request)
    if not active_project_id and projects:
        active_project_id = projects[0].get("id", 0)

    if not active_project_id:
        return render(request, "projects/monthly_allocations.html", {
            "projects": projects,
            "active_project_id": active_project_id,
            "month_start": month_start,
            "coes": [],
            "domains_map": {},
            "allocation_map": {},
            "capacity_map": {},
            "hours_available": HOURS_AVAILABLE_PER_MONTH,
            "weekly_map": {},
            "now": datetime.now(),
        })

    # Fetch COEs and domains
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT id, name FROM coes ORDER BY name")
            coes = dictfetchall(cur)
    except Exception:
        logger.exception("Error fetching COEs")
        coes = []
    coe_ids = [c["id"] for c in coes] if coes else []

    domains_map = {}
    if coe_ids:
        try:
            with connection.cursor() as cur:
                cur.execute("SELECT id, name, coe_id FROM domains WHERE coe_id IN %s ORDER BY name", [tuple(coe_ids)])
                doms = dictfetchall(cur)
            for d in doms:
                domains_map.setdefault(d["coe_id"], []).append({"id": d["id"], "name": d["name"]})
        except Exception:
            logger.exception("Error fetching domains")
            domains_map = {}

    # fetch allocation_items for this project/billing_start (canonical)
    allocation_map = {}
    capacity_accumulator = {}
    allocation_ids = []
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT ai.id AS item_id,
                       ai.allocation_id,
                       ai.coe_id,
                       ai.domain_id,
                       ai.user_ldap,
                       u.username AS username,
                       u.email AS email,
                       COALESCE(ai.total_hours,0) as total_hours
                FROM allocation_items ai
                JOIN allocations a ON ai.allocation_id = a.id
                LEFT JOIN users u ON ai.user_id = u.id
                WHERE ai.project_id = %s
                  AND a.month_start = %s
                ORDER BY ai.coe_id
            """, [active_project_id, month_start])
            items = dictfetchall(cur)

        for it in items:
            coe_id = it.get("coe_id") or 0
            ldap_val = (it.get("user_ldap") or "").strip()
            try:
                total_hours = round(float(it.get("total_hours") or 0.0), 2)
            except Exception:
                total_hours = 0.0

            allocation_map.setdefault(coe_id, []).append({
                "item_id": it.get("item_id"),
                "allocation_id": it.get("allocation_id"),
                "coe_id": coe_id,
                "domain_id": it.get("domain_id"),
                "user_ldap": ldap_val,
                "username": it.get("username"),
                "email": it.get("email"),
                "total_hours": total_hours,
                "w1": 0, "w2": 0, "w3": 0, "w4": 0,
                "s1": "", "s2": "", "s3": "", "s4": ""
            })
            if ldap_val:
                key = ldap_val.lower()
                capacity_accumulator[key] = round(capacity_accumulator.get(key, 0.0) + total_hours, 2)
            aid = it.get("allocation_id")
            if aid and aid not in allocation_ids:
                allocation_ids.append(aid)

    except Exception:
        logger.exception("Error fetching allocation_items")
        allocation_map = {}
        capacity_accumulator = {}
        allocation_ids = []

    # weekly allocations attach (unchanged)
    weekly_map = {}
    if allocation_ids:
        try:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT allocation_id, week_number, percent, status
                    FROM weekly_allocations
                    WHERE allocation_id IN %s
                """, [tuple(allocation_ids)])
                for r in dictfetchall(cur):
                    alloc = r["allocation_id"]
                    wk = int(r["week_number"])
                    weekly_map.setdefault(alloc, {})[wk] = {
                        "percent": float(r["percent"] or 0.0),
                        "status": (r.get("status") or "")
                    }
        except Exception:
            logger.exception("Error fetching weekly_allocations")
            weekly_map = {}

        for coe_id, items in allocation_map.items():
            for it in items:
                aid = it["allocation_id"]
                wk = weekly_map.get(aid, {})
                it["w1"] = wk.get(1, {}).get("percent", 0)
                it["w2"] = wk.get(2, {}).get("percent", 0)
                it["w3"] = wk.get(3, {}).get("percent", 0)
                it["w4"] = wk.get(4, {}).get("percent", 0)
                it["s1"] = wk.get(1, {}).get("status", "")
                it["s2"] = wk.get(2, {}).get("status", "")
                it["s3"] = wk.get(3, {}).get("status", "")
                it["s4"] = wk.get(4, {}).get("status", "")

    capacity_map = {}
    for ldap_key, allocated in capacity_accumulator.items():
        remaining = round(max(0.0, float(HOURS_AVAILABLE_PER_MONTH) - float(allocated)), 2)
        capacity_map[ldap_key] = {"allocated": round(float(allocated), 2), "remaining": remaining}

    # ensure every user in allocation_items has an entry in capacity_map
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT COALESCE(ai.user_ldap, '') as user_ldap
                FROM allocation_items ai
                JOIN allocations a ON ai.allocation_id = a.id
                WHERE ai.project_id = %s AND a.month_start = %s
            """, [active_project_id, month_start])
            for row in cur.fetchall():
                val = row[0] or ""
                key = val.strip().lower()
                if key and key not in capacity_map:
                    capacity_map[key] = {"allocated": 0, "remaining": HOURS_AVAILABLE_PER_MONTH}
    except Exception:
        pass

    return render(request, "projects/monthly_allocations.html", {
        "projects": projects,
        "active_project_id": active_project_id,
        "month_start": month_start,
        "coes": coes,
        "domains_map": domains_map,
        "allocation_map": allocation_map,
        "capacity_map": capacity_map,
        "hours_available": HOURS_AVAILABLE_PER_MONTH,
        "weekly_map": weekly_map,
        "now": datetime.now(),
    })


# Implement _get_month_hours_limit used above
def _get_month_hours_limit(year, month):
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT max_hours FROM monthly_hours_limit WHERE year = %s AND month = %s LIMIT 1", (int(year), int(month)))
            row = cur.fetchone()
            if row and row[0] is not None:
                return float(row[0])
    except Exception:
        logger.exception("_get_month_hours_limit failed")
    return float(HOURS_AVAILABLE_PER_MONTH)

@require_GET
def get_applicable_ioms(request):
    session_ldap = request.session.get("ldap_username")
    session_cn = request.session.get("cn")

    def _cn_to_creator(cn):
        if not cn: return ""
        parts = str(cn).strip().split()
        if len(parts) >= 2:
            return " ".join(parts[1:]) + " " + parts[0]
        return str(cn).strip()

    creator_candidates = []
    if session_cn:
        conv = _cn_to_creator(session_cn).strip()
        if conv and conv.lower() not in [c.lower() for c in creator_candidates]:
            creator_candidates.append(conv)
        s = session_cn.strip()
        if s and s.lower() not in [c.lower() for c in creator_candidates]:
            creator_candidates.append(s)
    if session_ldap:
        sld = session_ldap.strip()
        if sld and sld.lower() not in [c.lower() for c in creator_candidates]:
            creator_candidates.append(sld)
    creator_lower_vals = [c.lower() for c in creator_candidates]

    project_id = request.GET.get("project_id")
    try:
        year = int(request.GET.get("year") or datetime.now().year)
        month = int(request.GET.get("month") or datetime.now().month)
    except Exception:
        return HttpResponseBadRequest("Invalid year/month")

    _MONTH_MAP = {
        1: ("jan_fte", "jan_hours"),
        2: ("feb_fte", "feb_hours"),
        3: ("mar_fte", "mar_hours"),
        4: ("apr_fte", "apr_hours"),
        5: ("may_fte", "may_hours"),
        6: ("jun_fte", "jun_hours"),
        7: ("jul_fte", "jul_hours"),
        8: ("aug_fte", "aug_hours"),
        9: ("sep_fte", "sep_hours"),
        10: ("oct_fte", "oct_hours"),
        11: ("nov_fte", "nov_hours"),
        12: ("dec_fte", "dec_hours"),
    }
    fte_col, hrs_col = _MONTH_MAP.get(month, ("jan_fte", "jan_hours"))

    sql = f"""
        SELECT id, iom_id, department, site, `function`,
               {fte_col} as month_fte, {hrs_col} as month_hours,
               buyer_wbs_cc, seller_wbs_cc
        FROM prism_wbs
        WHERE year = %s
          AND ( ({fte_col} IS NOT NULL AND {fte_col} > 0) OR ({hrs_col} IS NOT NULL AND {hrs_col} > 0) )
    """
    params = [str(year)]
    if project_id:
        sql += " AND project_id = %s"
        params.append(project_id)
    if creator_lower_vals:
        placeholders = ",".join(["%s"] * len(creator_lower_vals))
        sql += f" AND LOWER(TRIM(creator)) IN ({placeholders})"
        params.extend(creator_lower_vals)

    sql += " ORDER BY iom_id LIMIT 500"

    try:
        with connection.cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall() or []
            cols = [c[0] for c in cur.description]
    except Exception as ex:
        logger.exception("get_applicable_ioms DB error: %s", ex)
        return JsonResponse({"ok": False, "error": str(ex)}, status=500)

    ioms = []
    # get month_limit (per-resource limit) once per request
    month_limit = _get_month_hours_limit(year, month)
    for r in rows:
        rec = dict(zip(cols, r))
        ioms.append({
            "id": rec.get("id"),
            "iom_id": rec.get("iom_id"),
            "department": rec.get("department"),
            "site": rec.get("site"),
            "function": rec.get("function"),
            "month_fte": float(rec.get("month_fte") or 0),
            "month_hours": float(rec.get("month_hours") or 0),
            "buyer_wbs_cc": rec.get("buyer_wbs_cc"),
            "seller_wbs_cc": rec.get("seller_wbs_cc"),
            "month_limit": float(month_limit),
        })
    return JsonResponse({"ok": True, "ioms": ioms})


# --- get_iom_details: fetch by id OR iom_id, compute remaining hours (from monthly_allocation_entries),
#     and remaining FTE = remaining_hours / month_limit(year,month) ---
@require_GET
def get_iom_details(request):
    iom_row_id = request.GET.get("iom_row_id")
    project_id = request.GET.get("project_id")
    try:
        year = int(request.GET.get("year") or datetime.now().year)
        month = int(request.GET.get("month") or datetime.now().month)
    except ValueError:
        return HttpResponseBadRequest("Invalid year/month")
    if not iom_row_id:
        return HttpResponseBadRequest("iom_row_id required")

    _MONTH_MAP = {
        1: ("jan_fte", "jan_hours"),
        2: ("feb_fte", "feb_hours"),
        3: ("mar_fte", "mar_hours"),
        4: ("apr_fte", "apr_hours"),
        5: ("may_fte", "may_hours"),
        6: ("jun_fte", "jun_hours"),
        7: ("jul_fte", "jul_hours"),
        8: ("aug_fte", "aug_hours"),
        9: ("sep_fte", "sep_hours"),
        10: ("oct_fte", "oct_hours"),
        11: ("nov_fte", "nov_hours"),
        12: ("dec_fte", "dec_hours"),
    }
    fte_col, hrs_col = _MONTH_MAP.get(month, ("jan_fte", "jan_hours"))

    try:
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT id, iom_id, project_id, department, site, `function`,
                       {fte_col} as month_fte, {hrs_col} as month_hours,
                       buyer_wbs_cc, seller_wbs_cc
                FROM prism_wbs
                WHERE id = %s OR iom_id = %s
                LIMIT 1
            """, [iom_row_id, iom_row_id])
            row = cur.fetchone()
            if not row:
                return JsonResponse({"ok": False, "error": "IOM not found"}, status=404)
            cols = [c[0] for c in cur.description]
            rec = dict(zip(cols, row))

            # Use canonical billing start for this year/month
            billing_start, billing_end = get_billing_period(year, month)

            # compute used hours for this IOM from monthly_allocation_entries using canonical month_start
            cur.execute("""
                SELECT COALESCE(SUM(total_hours),0) FROM monthly_allocation_entries
                WHERE project_id=%s AND iom_id=%s AND month_start=%s
            """, [project_id, rec.get("iom_id"), billing_start])
            used_hours = cur.fetchone()[0] or 0.0

    except Exception as ex:
        logger.exception("get_iom_details failed: %s", ex)
        return JsonResponse({"ok": False, "error": str(ex)}, status=500)

    month_hours = float(rec.get("month_hours") or 0.0)
    month_limit = _get_month_hours_limit(year, month)
    month_fte = round((month_hours / month_limit) if month_limit > 0 else 0.0, 2)

    remaining_hours = max(0.0, month_hours - float(used_hours))
    remaining_hours = round(remaining_hours, 2)
    remaining_fte = round((remaining_hours / month_limit) if month_limit > 0 else 0.0, 2)

    resp = {
        "ok": True,
        "iom": {
            "id": rec.get("id"),
            "iom_id": rec.get("iom_id"),
            "department": rec.get("department"),
            "site": rec.get("site"),
            "function": rec.get("function"),
            "month_fte": float(month_fte),
            "month_hours": float(round(month_hours, 2)),
            "total_fte": float(rec.get("total_fte") or 0),
            "total_hours": float(rec.get("total_hours") or 0),
            "buyer_wbs_cc": rec.get("buyer_wbs_cc"),
            "seller_wbs_cc": rec.get("seller_wbs_cc"),
            "remaining_hours": float(remaining_hours),
            "remaining_fte": float(remaining_fte),
            "month_limit": float(month_limit),
            "billing_start": billing_start,
            "billing_end": billing_end,
        }
    }

    return JsonResponse(resp)

def export_allocations(request):
    """
    Export allocations for an IOM and billing month. Accepts:
      - project_id, iom_id, and either month=YYYY-MM (preferred) OR month_start=YYYY-MM-DD
    """
    project_id = request.GET.get("project_id")
    iom_id = request.GET.get("iom_id")
    month_param = request.GET.get("month")  # YYYY-MM
    month_start_param = request.GET.get("month_start")

    if not (project_id and iom_id):
        return HttpResponseBadRequest("project_id and iom_id required")

    # resolve canonical billing_start based on month param or month_start
    billing_start = None
    if month_param:
        try:
            year, mon = map(int, month_param.split("-"))
            billing_start, billing_end = get_billing_period(year, mon)
        except Exception:
            logger.exception("export_allocations: invalid month param %s", month_param)
            from datetime import date
            billing_start = date.today().replace(day=1)
    elif month_start_param:
        try:
            dt = datetime.strptime(month_start_param, "%Y-%m-%d").date()
            # find billing period containing dt; if not found fallback to dt.replace(day=1)
            try:
                billing_start, billing_end = get_billing_period_for_date(dt)
            except Exception:
                billing_start = dt.replace(day=1)
        except Exception:
            billing_start = date.today().replace(day=1)
    else:
        today = date.today()
        billing_start, billing_end = get_billing_period(today.year, today.month)

    # fetch iom basic details
    iom = None
    with connection.cursor() as cur:
        cur.execute("""
            SELECT iom_id, department, buyer_wbs_cc, seller_wbs_cc, site, `function`, total_hours
            FROM prism_wbs
            WHERE project_id=%s AND iom_id=%s
            LIMIT 1
        """, [project_id, iom_id])
        row = cur.fetchone()
        if row:
            iom = {
                "iom_id": row[0],
                "department": row[1],
                "buyer_wbs_cc": row[2],
                "seller_wbs_cc": row[3],
                "site": row[4],
                "function": row[5],
                "total_hours": row[6],
            }

    with connection.cursor() as cur:
        cur.execute("""
            SELECT user_ldap, total_hours
            FROM monthly_allocation_entries
            WHERE project_id=%s AND iom_id=%s AND month_start=%s
        """, [project_id, iom_id, billing_start])
        allocations = cur.fetchall() or []

    # Build excel as before
    wb = Workbook()
    ws = wb.active
    ws.title = "Allocations"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    row_idx = 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    title_cell = ws.cell(row=row_idx, column=1, value="IOM Allocation Report")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.alignment = center
    title_cell.fill = fill_blue
    row_idx += 2

    if iom:
        details = [
            ("IOM ID", iom["iom_id"]),
            ("Department", iom["department"]),
            ("WBS", f'{iom["buyer_wbs_cc"] or "-"} / {iom["seller_wbs_cc"] or "-"}'),
            ("Site", iom["site"] or "-"),
            ("Function", iom["function"] or "-"),
            ("Total Hours", iom["total_hours"] or 0),
        ]
        for key, val in details:
            ws.cell(row=row_idx, column=1, value=key).font = bold_font
            ws.cell(row=row_idx, column=1).fill = fill_gray
            ws.cell(row=row_idx, column=1).border = border
            ws.cell(row=row_idx, column=2, value=val).font = normal_font
            ws.cell(row=row_idx, column=2).border = border
            row_idx += 1
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="Resource (LDAP)").font = header_font
    ws.cell(row=row_idx, column=1).fill = fill_blue
    ws.cell(row=row_idx, column=1).alignment = center
    ws.cell(row=row_idx, column=1).border = border
    ws.cell(row=row_idx, column=2, value="Total Hours").font = header_font
    ws.cell(row=row_idx, column=2).fill = fill_blue
    ws.cell(row=row_idx, column=2).alignment = center
    ws.cell(row=row_idx, column=2).border = border
    row_idx += 1

    for alloc in allocations:
        ws.cell(row=row_idx, column=1, value=alloc[0]).font = normal_font
        ws.cell(row=row_idx, column=1).alignment = left
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=2, value=float(alloc[1] or 0)).font = normal_font
        ws.cell(row=row_idx, column=2).alignment = right
        ws.cell(row=row_idx, column=2).border = border
        row_idx += 1

    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(idx)
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max_length + 4

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"allocations_{iom_id}_{billing_start}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_my_punches_pdf(request):
    """
    Export punches for the logged-in user for the billing cycle corresponding to the
    provided month (YYYY-MM). Uses get_billing_period(year, month) as source of truth.
    """
    import io
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa

    session_ldap = request.session.get("ldap_username") or request.session.get("user_email") or getattr(request.user, 'email', None)
    if not session_ldap:
        return HttpResponseBadRequest("Not authenticated")

    month_param = request.GET.get("month", date.today().strftime("%Y-%m"))
    try:
        year, month = map(int, month_param.split("-"))
    except Exception:
        year, month = date.today().year, date.today().month

    month_start, month_end = get_billing_period(year, month)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT up.allocation_id, mae.project_id, p.name as project_name, mae.iom_id, pw.department,
                   up.punch_date, up.week_number, up.actual_hours, up.wbs
            FROM user_punches up
            LEFT JOIN monthly_allocation_entries mae ON mae.id = up.allocation_id
            LEFT JOIN projects p ON mae.project_id = p.id
            LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
            WHERE up.user_ldap = %s
              AND up.punch_date BETWEEN %s AND %s
            ORDER BY up.punch_date, p.name
        """, [session_ldap, month_start, month_end])
        rows = dictfetchall(cur)

    html = render_to_string("projects/punches_pdf.html", {"rows": rows, "month": month_param, "user": session_ldap,
                                                          "billing_start": month_start, "billing_end": month_end})
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)
    result.seek(0)
    safe_user = session_ldap.replace("@", "_at_").replace(".", "_")
    filename = f"punches_{safe_user}_{month_param}.pdf"
    response = HttpResponse(result.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_my_punches_excel(request):
    """
    Export the logged-in user's punches for the billing cycle corresponding to `month` (YYYY-MM).
    Uses get_billing_period(year, month) as the single source of truth for start/end.
    """
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter

    session_ldap = request.session.get("ldap_username") or request.session.get("user_email") or getattr(request.user, 'email', None)
    if not session_ldap:
        return HttpResponseBadRequest("Not authenticated")

    month_param = request.GET.get("month", date.today().strftime("%Y-%m"))
    try:
        year, month = map(int, month_param.split("-"))
    except Exception:
        year, month = date.today().year, date.today().month

    month_start, month_end = get_billing_period(year, month)

    # fetch punches for user & billing window
    with connection.cursor() as cur:
        cur.execute("""
            SELECT up.allocation_id, mae.project_id, p.name as project_name, mae.iom_id, pw.department,
                   up.punch_date, up.week_number, up.actual_hours, up.wbs
            FROM user_punches up
            LEFT JOIN monthly_allocation_entries mae ON mae.id = up.allocation_id
            LEFT JOIN projects p ON mae.project_id = p.id
            LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
            WHERE up.user_ldap = %s
              AND up.punch_date BETWEEN %s AND %s
            ORDER BY up.punch_date, p.name
        """, [session_ldap, month_start, month_end])
        rows = dictfetchall(cur)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Punches {month_param}"

    headers = ["Date", "Project", "IOM", "Dept", "Week#", "Hours", "WBS"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = 20

    r = 2
    for rec in rows:
        # punch_date may already be a date object or a string depending on DB driver
        pd = rec.get('punch_date')
        if hasattr(pd, 'strftime'):
            pd_str = pd.strftime("%Y-%m-%d")
        else:
            pd_str = pd or ''
        ws.cell(row=r, column=1, value=pd_str)
        ws.cell(row=r, column=2, value=rec.get('project_name'))
        ws.cell(row=r, column=3, value=rec.get('iom_id'))
        ws.cell(row=r, column=4, value=rec.get('department'))
        ws.cell(row=r, column=5, value=rec.get('week_number'))
        ws.cell(row=r, column=6, value=float(rec.get('actual_hours') or 0))
        ws.cell(row=r, column=7, value=rec.get('wbs') or '')
        r += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_user = session_ldap.replace("@", "_at_").replace(".", "_")
    filename = f"punches_{safe_user}_{month_param}.xlsx"
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
